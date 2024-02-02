""" Fetches tariff XLSX files to find out rates for a given date and plan

Plans are must be a residential Time-Of-Use (TOU) plan in the following list:
* E-TOU-B - Peak 4-9pm on weekdays.  Weekends and Holidays are off-peak. No Baseline.
* E-TOU-C - Peak 4-9pm, with baseline usage rates (ie Tier 1 and Tier 2 pricing).  Same
* E-TOU-D - Peak 5-8pm.  Weekends and holidays are off-peak
"""
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import date, datetime
import re
import os
from collections import defaultdict

BASELINE_TERRITORY = 'W'
HEAT_SOURCE = 'B'  # "Basic Electricity" for baseline calculation
DAILY_BASELINE_KWH = 10.7
RATE_PAGE = 'https://www.pge.com/tariffs/'
CACHE = './cache/'
LINK_DICT = {}  # Tracks URLs of XLSX files for each time range
SUMMER = None # June-Sept

def parse_date(date_string):
    """ Parse date string from tariffs page """
    return datetime.strptime(date_string, '%b %d, %Y')


def parse_date_range(range_string):
    """ Returns tuple of dates indicating the date range in a string of a rate table link """
    since, through = re.match(r'.*\((.*) [–-] ([^\)]+)', range_string).groups()
    if through == 'Present':
        through = datetime.now()
    else:
        through = parse_date(through)
        through = through.replace(hour=23, minute=59, second=59)
    since = parse_date(since)
    return since, through


def get_link_dicts():
    """ Lists all date rate spreadsheets from https://www.pge.com/tariffs/electric.shtml
    as a dict with key tuple (from_date, to_date) and value is the hyperlink to the XLSX file """
    resp = requests.get(f"{RATE_PAGE}electric.shtml")
    soup = BeautifulSoup(resp.content, 'html.parser')
    lis = soup.find_all("ul")[10].find_all('li')  # list items
    link_dict = {}
    for li in lis:
        key = parse_date_range(li.find('a').text)
        value = f"{RATE_PAGE}{li.find('a')['href']}"
        link_dict[key] = value
    global LINK_DICT
    LINK_DICT = link_dict
    return link_dict


def get_rate_sheet(link: str):
    """ Given a date, select the right rate plan from link_dict and download to cache if it's not already there """
    local_path = f"{CACHE}{link.split('/')[-1]}"
    if not os.path.isfile(local_path):
        # File doesn't exist locally, so download it first
        r = requests.get(link, allow_redirects=True)
        with open(local_path, 'wb') as f:
            f.write(r.content)

    # Open file and parse rates
    return local_path
    # with open(local_path, 'r') as f:
    #     print(f)  # TODO

def parse_rate_sheet(filename):
    wb = load_workbook(filename=filename)
    rate_sheet = None
    baseline_sheet = None
    for sn in wb.sheetnames:
        if 'Res Inclu TOU' in sn:
            rate_sheet = sn
        elif 'ElecBaseline' in sn:
            baseline_sheet = sn

    # find basline allocation per day
    baseline_sheet = wb[baseline_sheet]
    baseline_dict = {}  # Summer and Winter keys
    # Winter range first
    baseline_range = baseline_sheet['A13:D36']

    for season, range in (('Winter', baseline_sheet['A13:D36']), ('Summer', baseline_sheet['F13:I36'])):
        for row in range:
            if row[0].value and 'Code' in row[0].value:
                code = row[0].value[-2]
                continue
            if code == HEAT_SOURCE and row[0].value == BASELINE_TERRITORY:
                baseline_dict[season] = row[1].value


    rate_sheet = wb[rate_sheet]
    rate_range = rate_sheet['A14:I25']
    rate_dict = defaultdict(dict)
    for row in rate_range:
        if row[0].value:  # Merged cells only have values in 1st cell.  Update our plan or use the old one
            if 'E-TOU-B' in row[0].value:
                plan = 'E-TOU-B'
            elif 'E-TOU-C' in row[0].value:
                plan = 'E-TOU-C'
            elif 'E-TOU-D' in row[0].value:
                plan = 'E-TOU-D'

        if row[5].value:
            season = row[5].value
        if row[6].value:
            period = row[6].value
        if season not in rate_dict[plan]:
            rate_dict[plan][season] = defaultdict(dict)
            rate_dict[plan][season]['baseline'] = baseline_dict[season]
        rate_dict[plan][season][period]['rate'] = row[7].value
        rate_dict[plan][season][period]['baseline_credit'] = row[8].value if row[8].value != '-' else 0
    return rate_dict


def get_rate(dt: datetime, usage: float = 0, bill_days: int = 0):
    """ Given a datetime of when power usage occurred, get the correct rate dict for each TOU plan.
    When baselines are used in rate calculations, pass in cumulative usage for this billing period
    along with days in this billing period
    """
    link_dict = LINK_DICT
    if not LINK_DICT:
        link_dict = get_link_dicts()
    link = None
    for k, v in link_dict.items():
        if k[0] <= dt <= k[1]:
            link = v
            break

    assert link is not None
    filename = get_rate_sheet(link)
    rates = parse_rate_sheet(filename)
    """
    defaultdict(<class 'dict'>, {'E-TOU-B': {'Summer': defaultdict(<class 'dict'>, {'baseline': 19.2, 'Peak': {'rate': 0.5342600000000001, 'baseline_credit': 0}, 'Off-Peak': {'rate': 0.41120000000000007, 'baseline_credit': 0}}), 'Winter': defaultdict(<class 'dict'>, {'baseline': 9.8, 'Peak': {'rate': 0.39762999999999993, 'baseline_credit': 0}, 'Off-Peak': {'rate': 0.35883, 'baseline_credit': 0}})}, 'E-TOU-C': {'Summer': defaultdict(<class 'dict'>, {'baseline': 19.2, 'Peak': {'rate': 0.53933, 'baseline_credit': -0.08851}, 'Off-Peak': {'rate': 0.45589, 'baseline_credit': -0.08851}}), 'Winter': defaultdict(<class 'dict'>, {'baseline': 9.8, 'Peak': {'rate': 0.43662, 'baseline_credit': -0.08851}, 'Off-Peak': {'rate': 0.40827, 'baseline_credit': -0.08851}})}, 'E-TOU-D': {'Summer': defaultdict(<class 'dict'>, {'baseline': 19.2, 'Peak': {'rate': 0.51778, 'baseline_credit': 0}, 'Off-Peak': {'rate': 0.38282, 'baseline_credit': 0}}), 'Winter': defaultdict(<class 'dict'>, {'baseline': 9.8, 'Peak': {'rate': 0.42818, 'baseline_credit': 0}, 'Off-Peak': {'rate': 0.38957, 'baseline_credit': 0}})}})
    """

    resp = {}
    season = 'Summer' if dt.month in (6, 7, 8, 9) else 'Winter'

    for plan, details in rates.items():
        if plan == 'E-TOU-B':
            period = 'Peak' if 16 <= dt.hour < 21 else 'Off-Peak'
            rate = details[season][period]['rate']
            resp[plan] = rate
        elif plan == 'E-TOU-C':
            period = 'Peak' if 16 <= dt.hour < 21 else 'Off-Peak'
            rate = details[season][period]['rate']
            if usage <= bill_days * details[season]['baseline']:
                rate += details[season][period]['baseline_credit']
            resp[plan] = rate
        elif plan == 'E-TOU-D':
            period = 'Peak' if 17 <= dt.hour < 20 else 'Off-Peak'
            rate = details[season][period]['rate']
            resp[plan] = rate
    return resp




if __name__ == '__main__':
    dt = datetime(2023,10,4,6,0)
    print(get_rate(dt))